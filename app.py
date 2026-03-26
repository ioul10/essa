import streamlit as st
import pandas as pd
import io
from extractor import PDFExtractor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# Configuration
st.set_page_config(page_title="Extracteur PDF Intelligent", page_icon="🧠", layout="wide")

# CSS personnalisé
st.markdown("""
<style>
    .step-box {
        background-color: #f0f2f6;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
    }
    .success-box {
        background-color: #d4edda;
        padding: 10px;
        border-radius: 5px;
        margin: 5px 0;
    }
</style>
""", unsafe_allow_html=True)

st.title("🧠 Extracteur PDF Intelligent vers Excel")
st.markdown("**Pipeline complet** : Extraction → Nettoyage → Filtrage → Mapping → Template")

# Initialisation
extractor = PDFExtractor()

# Upload
uploaded_file = st.file_uploader("📄 Charger votre PDF (Bilan, Facture, etc.)", type="pdf")

if uploaded_file:
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("📊 Pipeline de Traitement")
        
        # Étape 1 & 2
        with st.expander("✅ Étape 1-2: Extraction Texte & Tableaux", expanded=True):
            full_text, tables = extractor.extract_text(uploaded_file)
            st.write(f"**Texte extrait** : {len(full_text)} caractères")
            st.write(f"**Tableaux détectés** : {len(tables)}")
        
        # Étape 3
        with st.expander("✅ Étape 3: Nettoyage des Données", expanded=True):
            cleaned_tables = extractor.clean_data(tables)
            st.write(f"**Tableaux après nettoyage** : {len(cleaned_tables)}")
        
        # Étape 4
        with st.expander("✅ Étape 4: Filtrage Lignes Utiles", expanded=False):
            useful_tables = extractor.filter_useful_lines(cleaned_tables)
            st.write(f"**Tableaux utiles** : {len(useful_tables)}")
            
            if useful_tables:
                st.dataframe(pd.DataFrame(useful_tables[0][:5]))  # Aperçu
        
        # Étape 5
        with st.expander("✅ Étape 5: Extraction Montants", expanded=False):
            amounts = extractor.extract_amounts(full_text)
            st.write(f"**Montants trouvés** : {len(amounts)}")
            if amounts:
                st.write(f"Exemples : {amounts[:5]}")
        
        # Étape 6 & 7
        with st.expander("✅ Étape 6-7: Mapping & Template", expanded=True):
            key_values = extractor.extract_key_values(full_text)
            template_df = extractor.create_template_df()
            filled_df = extractor.fill_template(key_values, template_df)
            st.dataframe(filled_df)
    
    with col2:
        st.subheader("📥 Export Excel")
        
        # Création Excel
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Feuille 1: Données structurées
            filled_df.to_excel(writer, sheet_name='Données_Extraites', index=False)
            
            # Feuille 2: Tableaux bruts
            for idx, table in enumerate(useful_tables[:5]):  # Max 5 tableaux
                if len(table) > 1:
                    df_table = pd.DataFrame(table[1:], columns=table[0])
                    df_table.to_excel(writer, sheet_name=f'Tableau_{idx+1}', index=False)
            
            # Feuille 3: Texte brut
            text_df = pd.DataFrame({'Texte Extrait': [full_text]})
            text_df.to_excel(writer, sheet_name='Texte_Brut', index=False)
        
        # Stylisation
        from openpyxl import load_workbook
        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)
        
        for ws in wb.worksheets:
            # En-têtes
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Largeur auto
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = min(adjusted_width, 50)
        
        # Sauvegarde
        wb.save(excel_buffer)
        excel_bytes = excel_buffer.getvalue()
        
        st.download_button(
            label="📥 Télécharger Excel Complet",
            data=excel_bytes,
            file_name=f"extraction_{uploaded_file.name.replace('.pdf', '')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Stats
        st.markdown("### 📈 Statistiques")
        st.metric("Champs Mappés", len([v for v in key_values.values() if v]))
        st.metric("Tableaux Exportés", min(len(useful_tables), 5))
        st.metric("Montants Extraits", len(amounts))

else:
    st.info("👆 Chargez un fichier PDF pour commencer l'extraction")
    st.markdown("""
    ### 💡 Types de documents supportés
    - ✅ Bilans comptables
    - ✅ Factures fournisseurs
    - ✅ Comptes de résultat
    - ✅ Tableaux financiers
    - ✅ Documents structurés avec texte
    """)
